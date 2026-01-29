"""导出服务

提供批改结果的多种导出功能：
- 带批注的学生作答图片 (ZIP)
- 班级统计 Excel
- LLM 智能 Excel 生成
"""

import io
import json
import logging
import zipfile
import base64
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.services.annotation_renderer import AnnotationRenderer, RenderConfig
from src.models.annotation import VisualAnnotation, PageAnnotations, BoundingBox, AnnotationType


logger = logging.getLogger(__name__)


@dataclass
class ExportConfig:
    """导出配置"""

    include_original: bool = False  # 是否包含原始图片
    image_format: str = "PNG"  # 图片格式
    image_quality: int = 95  # JPEG 质量
    excel_template: Optional[str] = None  # Excel 模板路径


class AnnotatedImageExporter:
    """带批注图片导出器"""

    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()
        self.renderer = AnnotationRenderer(
            RenderConfig(
                output_format=self.config.image_format,
                output_quality=self.config.image_quality,
            )
        )

    def _build_annotations_from_results(
        self,
        question_results: List[Dict[str, Any]],
        page_index: int,
    ) -> List[VisualAnnotation]:
        """从批改结果构建批注列表"""
        annotations = []

        for q in question_results:
            q_page_indices = q.get("page_indices") or q.get("pageIndices") or []
            if page_index not in q_page_indices and q_page_indices:
                continue

            # 从 questionResult 中提取批注
            raw_annotations = q.get("annotations") or []
            for ann in raw_annotations:
                if ann.get("page_index") == page_index or not ann.get("page_index"):
                    try:
                        bbox = ann.get("bounding_box") or ann.get("boundingBox") or {}
                        if bbox:
                            annotations.append(
                                VisualAnnotation(
                                    annotation_type=AnnotationType(ann.get("type", "comment")),
                                    bounding_box=BoundingBox(
                                        x_min=bbox.get("x_min", 0),
                                        y_min=bbox.get("y_min", 0),
                                        x_max=bbox.get("x_max", 0),
                                        y_max=bbox.get("y_max", 0),
                                    ),
                                    text=ann.get("text"),
                                    color=ann.get("color", "#FF0000"),
                                )
                            )
                    except Exception as e:
                        logger.warning(f"解析批注失败: {e}")

            # 从 scoringPointResults 中提取 errorRegion
            for spr in q.get("scoringPointResults") or q.get("scoring_point_results") or []:
                error_region = spr.get("errorRegion") or spr.get("error_region")
                if error_region:
                    awarded = spr.get("awarded", 0)
                    max_points = spr.get("maxPoints") or spr.get("max_points", 0)
                    is_correct = awarded >= max_points if max_points > 0 else awarded > 0

                    annotations.append(
                        VisualAnnotation(
                            annotation_type=(
                                AnnotationType.CORRECT_CHECK
                                if is_correct
                                else AnnotationType.ERROR_CIRCLE
                            ),
                            bounding_box=BoundingBox(
                                x_min=error_region.get("x_min", 0),
                                y_min=error_region.get("y_min", 0),
                                x_max=error_region.get("x_max", 0),
                                y_max=error_region.get("y_max", 0),
                            ),
                            color="#22C55E" if is_correct else "#EF4444",
                        )
                    )

            # 从 steps 中提取步骤标注
            for step in q.get("steps") or []:
                step_region = step.get("step_region") or step.get("stepRegion")
                if step_region:
                    is_correct = step.get("is_correct", False)
                    annotations.append(
                        VisualAnnotation(
                            annotation_type=(
                                AnnotationType.STEP_CHECK
                                if is_correct
                                else AnnotationType.STEP_CROSS
                            ),
                            bounding_box=BoundingBox(
                                x_min=step_region.get("x_min", 0),
                                y_min=step_region.get("y_min", 0),
                                x_max=step_region.get("x_max", 0),
                                y_max=step_region.get("y_max", 0),
                            ),
                            color="#22C55E" if is_correct else "#EF4444",
                        )
                    )

            # 答案区域标注
            answer_region = q.get("answerRegion") or q.get("answer_region")
            if answer_region:
                score = q.get("score", 0)
                max_score = q.get("maxScore") or q.get("max_score", 0)
                # 添加分数标注
                annotations.append(
                    VisualAnnotation(
                        annotation_type=AnnotationType.SCORE,
                        bounding_box=BoundingBox(
                            x_min=answer_region.get("x_max", 0) + 0.01,
                            y_min=answer_region.get("y_min", 0),
                            x_max=min(answer_region.get("x_max", 0) + 0.08, 1.0),
                            y_max=answer_region.get("y_min", 0) + 0.04,
                        ),
                        text=f"{score}/{max_score}",
                        color="#3B82F6",
                    )
                )

        return annotations

    def render_student_pages(
        self,
        student_result: Dict[str, Any],
        images: List[bytes],
        start_page: int = 0,
        end_page: Optional[int] = None,
    ) -> List[Tuple[int, bytes]]:
        """
        渲染学生的所有页面

        Returns:
            List of (page_index, rendered_image_bytes)
        """
        rendered = []
        question_results = (
            student_result.get("questionResults") or student_result.get("question_results") or []
        )

        if end_page is None:
            end_page = len(images) - 1

        for page_idx in range(start_page, min(end_page + 1, len(images))):
            image_data = images[page_idx]

            # 构建该页的批注
            annotations = self._build_annotations_from_results(question_results, page_idx)

            if annotations:
                # 渲染批注
                page_annotations = PageAnnotations(page_index=page_idx, annotations=annotations)
                rendered_image = self.renderer.render_page(image_data, page_annotations)
            else:
                rendered_image = image_data

            rendered.append((page_idx, rendered_image))

        return rendered

    def export_to_zip(
        self,
        student_results: List[Dict[str, Any]],
        images: List[bytes],
        batch_id: str,
    ) -> bytes:
        """
        导出所有学生的带批注图片为 ZIP

        Args:
            student_results: 学生批改结果列表
            images: 原始图片列表
            batch_id: 批次 ID

        Returns:
            ZIP 文件字节
        """
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for student in student_results:
                student_name = (
                    student.get("studentName")
                    or student.get("student_name")
                    or student.get("student_key")
                    or "Unknown"
                )
                start_page = student.get("startPage") or student.get("start_page") or 0
                end_page = student.get("endPage") or student.get("end_page")

                # 渲染该学生的页面
                rendered_pages = self.render_student_pages(student, images, start_page, end_page)

                # 添加到 ZIP
                for page_idx, image_bytes in rendered_pages:
                    filename = f"{student_name}/page_{page_idx + 1}.png"
                    zf.writestr(filename, image_bytes)

                # 如果需要原始图片
                if self.config.include_original:
                    for page_idx in range(start_page, (end_page or len(images) - 1) + 1):
                        if page_idx < len(images):
                            filename = f"{student_name}/original_page_{page_idx + 1}.png"
                            zf.writestr(filename, images[page_idx])

            # 添加元数据
            metadata = {
                "batch_id": batch_id,
                "export_time": datetime.now().isoformat(),
                "student_count": len(student_results),
                "total_pages": len(images),
            }
            zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))

        zip_buffer.seek(0)
        return zip_buffer.getvalue()


class ExcelExporter:
    """Excel 导出器"""

    # 默认列配置
    DEFAULT_COLUMNS = [
        {"key": "rank", "header": "排名", "width": 8},
        {"key": "studentName", "header": "学生姓名", "width": 15},
        {"key": "score", "header": "得分", "width": 10},
        {"key": "maxScore", "header": "满分", "width": 10},
        {"key": "percentage", "header": "得分率", "width": 12},
        {"key": "pageRange", "header": "页码范围", "width": 12},
    ]

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()

    def _get_nested_value(self, obj: Dict, key: str) -> Any:
        """获取嵌套字典值，支持点号分隔"""
        keys = key.split(".")
        value = obj
        for k in keys:
            if isinstance(value, dict):
                value = value.get(k)
            else:
                return None
        return value

    def export_basic(
        self,
        student_results: List[Dict[str, Any]],
        class_report: Optional[Dict[str, Any]] = None,
        columns: Optional[List[Dict[str, str]]] = None,
    ) -> bytes:
        """
        导出基础 Excel 统计

        Args:
            student_results: 学生结果列表
            class_report: 班级报告（可选）
            columns: 自定义列配置

        Returns:
            Excel 文件字节
        """
        wb = Workbook()

        # ===== Sheet 1: 学生成绩 =====
        ws = wb.active
        ws.title = "学生成绩"

        cols = columns or self.DEFAULT_COLUMNS

        # 写入表头
        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["header"])
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col.get("width", 12)

        # 写入数据
        sorted_results = sorted(
            student_results,
            key=lambda x: x.get("score", 0),
            reverse=True,
        )

        for row_idx, student in enumerate(sorted_results, 2):
            for col_idx, col in enumerate(cols, 1):
                key = col["key"]

                if key == "rank":
                    value = row_idx - 1
                elif key == "percentage":
                    score = student.get("score", 0)
                    max_score = student.get("maxScore") or student.get("max_score", 100)
                    value = f"{(score / max_score * 100):.1f}%" if max_score > 0 else "0%"
                else:
                    value = self._get_nested_value(student, key)
                    if value is None:
                        # 尝试 snake_case
                        snake_key = "".join(
                            f"_{c.lower()}" if c.isupper() else c for c in key
                        ).lstrip("_")
                        value = self._get_nested_value(student, snake_key)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ===== Sheet 2: 题目统计 =====
        ws2 = wb.create_sheet("题目统计")

        # 收集所有题目
        question_stats: Dict[str, Dict[str, Any]] = {}
        for student in student_results:
            for q in student.get("questionResults") or student.get("question_results") or []:
                qid = q.get("questionId") or q.get("question_id", "")
                if qid not in question_stats:
                    question_stats[qid] = {
                        "questionId": qid,
                        "maxScore": q.get("maxScore") or q.get("max_score", 0),
                        "scores": [],
                    }
                question_stats[qid]["scores"].append(q.get("score", 0))

        # 表头
        q_headers = ["题号", "满分", "平均分", "得分率", "最高分", "最低分", "答对人数"]
        for col_idx, header in enumerate(q_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        # 数据
        for row_idx, (qid, stats) in enumerate(sorted(question_stats.items()), 2):
            scores = stats["scores"]
            max_score = stats["maxScore"]
            avg_score = sum(scores) / len(scores) if scores else 0

            values = [
                qid,
                max_score,
                round(avg_score, 2),
                f"{(avg_score / max_score * 100):.1f}%" if max_score > 0 else "0%",
                max(scores) if scores else 0,
                min(scores) if scores else 0,
                sum(1 for s in scores if s >= max_score),
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center")

        # ===== Sheet 3: 班级报告 =====
        if class_report:
            ws3 = wb.create_sheet("班级报告")

            report_items = [
                ("总人数", class_report.get("totalStudents") or class_report.get("total_students")),
                ("平均分", class_report.get("averageScore") or class_report.get("average_score")),
                (
                    "平均得分率",
                    f"{(class_report.get('averagePercentage') or class_report.get('average_percentage', 0)):.1f}%",
                ),
                (
                    "及格率",
                    f"{(class_report.get('passRate') or class_report.get('pass_rate', 0)):.1f}%",
                ),
                ("生成时间", class_report.get("generatedAt") or class_report.get("generated_at")),
            ]

            for row_idx, (label, value) in enumerate(report_items, 1):
                ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws3.cell(row=row_idx, column=2, value=value)

            # 薄弱知识点
            weak_points = class_report.get("weakPoints") or class_report.get("weak_points") or []
            if weak_points:
                ws3.cell(row=len(report_items) + 2, column=1, value="薄弱知识点").font = Font(
                    bold=True
                )
                for i, wp in enumerate(weak_points):
                    ws3.cell(
                        row=len(report_items) + 3 + i,
                        column=1,
                        value=f"{wp.get('pointId', '')} - {wp.get('description', '')} ({wp.get('masteryRatio', 0):.0%})",
                    )

        # 输出
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def merge_with_template(
        self,
        template_bytes: bytes,
        student_results: List[Dict[str, Any]],
        mapping: Dict[str, str],
    ) -> bytes:
        """
        将数据合并到已有 Excel 模板

        Args:
            template_bytes: 模板 Excel 字节
            student_results: 学生结果
            mapping: 列映射 {模板列名: 数据字段名}

        Returns:
            合并后的 Excel 字节
        """
        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active

        # 找到表头行
        header_row = 1
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers[str(cell_value).strip()] = col

        # 找到数据起始行
        data_start_row = header_row + 1

        # 按学生姓名匹配并填充
        for student in student_results:
            student_name = (
                student.get("studentName")
                or student.get("student_name")
                or student.get("student_key")
            )

            # 查找学生行
            target_row = None
            name_col = headers.get("姓名") or headers.get("学生姓名") or headers.get("name")

            if name_col:
                for row in range(data_start_row, ws.max_row + 1):
                    if ws.cell(row=row, column=name_col).value == student_name:
                        target_row = row
                        break

            if target_row is None:
                # 新增行
                target_row = ws.max_row + 1
                if name_col:
                    ws.cell(row=target_row, column=name_col, value=student_name)

            # 填充映射的列
            for template_col, data_field in mapping.items():
                if template_col in headers:
                    col_idx = headers[template_col]
                    value = self._get_nested_value(student, data_field)
                    if value is not None:
                        ws.cell(row=target_row, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class SmartExcelGenerator:
    """LLM 智能 Excel 生成器"""

    def __init__(self, llm_client=None):
        self.llm_client = llm_client
        self.excel_exporter = ExcelExporter()

    async def generate_from_prompt(
        self,
        student_results: List[Dict[str, Any]],
        class_report: Optional[Dict[str, Any]],
        user_prompt: str,
        template_bytes: Optional[bytes] = None,
    ) -> Tuple[bytes, str]:
        """
        根据用户提示生成 Excel

        Args:
            student_results: 学生结果
            class_report: 班级报告
            user_prompt: 用户描述的格式需求
            template_bytes: 可选的模板 Excel

        Returns:
            (Excel 字节, LLM 解释说明)
        """
        # 构建可用字段列表
        available_fields = self._extract_available_fields(student_results)

        # 构建 LLM 提示
        system_prompt = """你是一个 Excel 报表生成助手。用户会描述他们想要的报表格式，你需要：
1. 分析用户需求
2. 从可用字段中选择合适的列
3. 返回 JSON 格式的列配置

可用字段：
{fields}

返回格式：
```json
{{
  "columns": [
    {{"key": "字段名", "header": "列标题", "width": 列宽}},
    ...
  ],
  "explanation": "解释说明"
}}
```"""

        if template_bytes:
            # 分析模板结构
            template_info = self._analyze_template(template_bytes)
            system_prompt += f"\n\n用户提供了模板，模板列：{template_info}"

        # 调用 LLM
        if self.llm_client:
            try:
                response = await self.llm_client.generate(
                    system_prompt.format(fields=json.dumps(available_fields, ensure_ascii=False)),
                    user_prompt,
                )

                # 解析 LLM 响应
                config = self._parse_llm_response(response)

                if template_bytes and config.get("mapping"):
                    # 合并到模板
                    excel_bytes = self.excel_exporter.merge_with_template(
                        template_bytes,
                        student_results,
                        config["mapping"],
                    )
                else:
                    # 生成新 Excel
                    excel_bytes = self.excel_exporter.export_basic(
                        student_results,
                        class_report,
                        config.get("columns"),
                    )

                return excel_bytes, config.get("explanation", "已根据您的需求生成报表")

            except Exception as e:
                logger.error(f"LLM 生成失败: {e}")

        # 回退到默认导出
        excel_bytes = self.excel_exporter.export_basic(student_results, class_report)
        return excel_bytes, "已使用默认格式生成报表"

    def _extract_available_fields(self, student_results: List[Dict[str, Any]]) -> List[str]:
        """提取可用字段"""
        fields = set()

        if student_results:
            sample = student_results[0]
            self._collect_fields(sample, "", fields)

        return sorted(list(fields))

    def _collect_fields(self, obj: Any, prefix: str, fields: set):
        """递归收集字段"""
        if isinstance(obj, dict):
            for key, value in obj.items():
                full_key = f"{prefix}.{key}" if prefix else key
                fields.add(full_key)
                if isinstance(value, dict):
                    self._collect_fields(value, full_key, fields)
                elif isinstance(value, list) and value and isinstance(value[0], dict):
                    self._collect_fields(value[0], f"{full_key}[]", fields)

    def _analyze_template(self, template_bytes: bytes) -> List[str]:
        """分析模板列"""
        try:
            wb = load_workbook(io.BytesIO(template_bytes))
            ws = wb.active
            headers = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=1, column=col).value
                if value:
                    headers.append(str(value))
            return headers
        except Exception as e:
            logger.warning(f"分析模板失败: {e}")
            return []

    def _parse_llm_response(self, response: str) -> Dict[str, Any]:
        """解析 LLM 响应"""
        try:
            # 提取 JSON
            import re

            json_match = re.search(r"```json\s*(.*?)\s*```", response, re.DOTALL)
            if json_match:
                return json.loads(json_match.group(1))

            # 尝试直接解析
            return json.loads(response)
        except Exception as e:
            logger.warning(f"解析 LLM 响应失败: {e}")
            return {}


# 导出
__all__ = [
    "ExportConfig",
    "AnnotatedImageExporter",
    "ExcelExporter",
    "SmartExcelGenerator",
]
